"""
教师管理 API。鉴权：仅 **HttpOnly 签名 Cookie** `teacher_session`（无教师端 Bearer，设计 §3.1）。

登录与 bootstrap 成功后会 **Set-Cookie: teacher_session=...**（HMAC-SHA256 签名，见 `app.services.teacher_session`）。

名单 `POST /v1/admin/roster/import` 支持 `multipart` 上传文件（`file` 字段，**.csv、.xlsx 或 .json**）或
`Content-Type: application/json`。**CSV / xlsx** 首行表头须含 **学号**、**姓名**、**班级**（UTF-8）；**JSON** 可为 `rows[].studentNo/fullName/className`（`className` 可选，另支持别名字段 `班级`）。
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from collections.abc import Iterator
from datetime import datetime, timezone

from fastapi import (
    APIRouter,
    File,
    Header,
    HTTPException,
    Request,
    Response,
    UploadFile,
    status,
)
from pydantic import AliasChoices, BaseModel, ConfigDict, Field
from sqlalchemy import func, select, update

from ..config import settings
from ..db.models import AdminAudit, AdminConfig, Class, RosterEntry, Student
from ..deps import CurrentTeacher, DBSession, require_bootstrap_token
from ..services.crypto import decrypt_password
from ..services.teacher_session import create_teacher_session_value

try:
    from passlib.context import CryptContext
except ImportError:  # pragma: no cover
    CryptContext = None  # type: ignore[misc, assignment]

router = APIRouter(prefix="/v1/admin", tags=["admin"])

_pwd = CryptContext(schemes=["bcrypt"], deprecated="auto") if CryptContext else None


class PasswordBody(BaseModel):
    model_config = {"populate_by_name": True}

    password: str = Field(min_length=1, description="管理员密码", alias="password")


def _set_teacher_cookie(response: Response) -> None:
    val = create_teacher_session_value()
    response.set_cookie(
        key="teacher_session",
        value=val,
        httponly=True,
        max_age=7 * 24 * 3600,
        samesite="lax",
        secure=settings.admin_cookie_secure,
        path="/",
    )


@router.post("/bootstrap", status_code=status.HTTP_201_CREATED)
async def bootstrap(
    body: PasswordBody,
    response: Response,
    db: DBSession,
    x_admin_bootstrap_token: str | None = Header(default=None, alias="X-Admin-Bootstrap-Token"),
) -> dict:
    """
    首次写入 `admin_config`。**若表已有配置则 403。**

    当环境配置 `admin_bootstrap_token` 时，须同时请求头
    `X-Admin-Bootstrap-Token: <同值>`，否则 403。
    """
    require_bootstrap_token(x_admin_bootstrap_token)
    r = await db.execute(select(AdminConfig).where(AdminConfig.id == 1))
    if r.scalar_one_or_none() is not None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="admin already bootstrapped",
        )
    if _pwd is None:
        raise RuntimeError("passlib not installed")
    h = _pwd.hash(body.password)
    db.add(AdminConfig(id=1, password_hash=h))
    _set_teacher_cookie(response)
    return {"ok": True}


@router.post("/login")
async def login(body: PasswordBody, response: Response, db: DBSession) -> dict:
    """校验 bcrypt 后设置 `teacher_session`（HttpOnly、签名，见模块文档字符串）。"""
    r = await db.execute(select(AdminConfig).where(AdminConfig.id == 1))
    row = r.scalar_one_or_none()
    if row is None or _pwd is None or not _pwd.verify(body.password, row.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="invalid credentials",
        )
    await db.execute(
        update(AdminConfig).where(AdminConfig.id == 1).values(updated_at=func.now())
    )
    _set_teacher_cookie(response)
    return {"ok": True}


@router.get("/me")
async def me(teacher: CurrentTeacher) -> dict:
    return {"ok": True, "role": "admin", "sub": teacher}


# --- 任务 4：名单导入 ---


class RosterItem(BaseModel):
    model_config = ConfigDict(populate_by_name=True, str_strip_whitespace=True)

    student_no: str = Field(max_length=64, description="学号，唯一", alias="studentNo")
    full_name: str = Field(
        max_length=255, description="姓名，首登须与名单完全对照", alias="fullName"
    )
    class_name: str | None = Field(
        default=None,
        max_length=128,
        description="班级名称；可空",
        validation_alias=AliasChoices("className", "班级"),
    )


class RosterJsonBody(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    rows: list[RosterItem]


def _iter_csv_rows(data: str) -> Iterator[tuple[str, str, str | None]]:
    f = io.StringIO(data)
    r = csv.DictReader(f)
    if not r.fieldnames:
        return
    fieldnames = [h.strip() for h in (r.fieldnames or []) if h and h.strip()]
    for need in ("学号", "姓名", "班级"):
        if need not in fieldnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="csv_headers_missing: 表头需含列「学号」「姓名」「班级」",
            )
    for row in r:
        a = (row.get("学号") or "").strip()
        b = (row.get("姓名") or "").strip()
        c = (row.get("班级") or "").strip() or None
        if not a and not b:
            continue
        if not a or not b:
            continue
        yield a, b, c


def _rows_from_json_raw(raw: object) -> list[tuple[str, str, str | None]]:
    if isinstance(raw, list):
        items = [RosterItem.model_validate(x) for x in raw]
    else:
        body = RosterJsonBody.model_validate(raw)
        items = body.rows
    return [
        (
            i.student_no.strip(),
            i.full_name.strip(),
            (i.class_name.strip() if (i.class_name or "").strip() else None),
        )
        for i in items
    ]


def _load_rows_xlsx(content: bytes) -> list[tuple[str, str, str | None]]:
    """首行表头须含列名 **学号**、**姓名**、**班级**（列名去首尾空格）。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="xlsx support unavailable",
        ) from e
    bio = io.BytesIO(content)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return []

        def cell_str(v: object) -> str:
            if v is None:
                return ""
            return str(v).strip()

        hlist = [cell_str(c) for c in header]
        for label, err in (
            ("学号", "学号"),
            ("姓名", "姓名"),
            ("班级", "班级"),
        ):
            if label not in hlist:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"xlsx_headers_missing: 首行表头需含列「{err}」",
                )
        i_no = hlist.index("学号")
        i_name = hlist.index("姓名")
        i_class = hlist.index("班级")
        out: list[tuple[str, str, str | None]] = []
        for row in it:
            if row is None:
                continue
            a = cell_str(row[i_no]) if i_no < len(row) else ""
            b = cell_str(row[i_name]) if i_name < len(row) else ""
            c_raw = cell_str(row[i_class]) if i_class < len(row) else ""
            c = c_raw or None
            if not a and not b:
                continue
            if not a or not b:
                continue
            out.append((a, b, c))
        return out
    finally:
        wb.close()


def _load_rows_from_file(content: bytes, filename: str) -> list[tuple[str, str, str | None]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _load_rows_xlsx(content)
    as_json = name.endswith(".json")
    if not as_json and content and content[0:1] in (b"{", b"["):
        as_json = True
    if as_json:
        try:
            raw = json.loads(content.decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="invalid json",
            ) from e
        return _rows_from_json_raw(raw)
    text = content.decode("utf-8-sig")
    return list(_iter_csv_rows(text))


async def _get_or_create_class_id(
    db: DBSession, class_name: str | None
) -> uuid.UUID | None:
    if not (class_name or "").strip():
        return None
    name = (class_name or "").strip()[:128]
    r = await db.execute(select(Class).where(Class.name == name))
    ex = r.scalar_one_or_none()
    if ex is not None:
        return ex.id
    c = Class(name=name)
    db.add(c)
    await db.flush()
    return c.id


async def _upsert_roster_row(
    db: DBSession, student_no: str, full_name: str, class_id: uuid.UUID | None
) -> None:
    st_r = await db.execute(select(Student).where(Student.student_no == student_no))
    student = st_r.scalar_one_or_none()
    e_r = await db.execute(
        select(RosterEntry).where(RosterEntry.student_no == student_no)
    )
    entry = e_r.scalar_one_or_none()
    if entry is None:
        db.add(
            RosterEntry(
                student_no=student_no,
                full_name=full_name,
                status="bound" if student is not None else "pending",
                deleted_at=None,
                student_id=student.id if student is not None else None,
                class_id=class_id,
            )
        )
        if student is not None and class_id is not None:
            student.class_id = class_id
        return
    if entry.deleted_at is not None:
        entry.deleted_at = None
    entry.full_name = full_name
    entry.class_id = class_id
    if student is not None:
        entry.student_id = student.id
        entry.status = "bound"
        if class_id is not None:
            student.class_id = class_id
    else:
        entry.student_id = None
        entry.status = "pending"


@router.post("/roster/import", status_code=status.HTTP_200_OK)
async def import_roster(
    _t: CurrentTeacher,
    db: DBSession,
    request: Request,
    file: UploadFile | None = File(default=None, description="CSV 或 JSON，UTF-8，字段名 studentNo, fullName"),
) -> dict:
    """
    导入或更新 `roster_entries`；尚无 `students` 学号时为 **pending**；已存在学号时 **bound** 并关联 `student_id`。
    - **JSON**：`Content-Type: application/json`，body 为 `{"rows":[{"studentNo","fullName"},...]}` 或 **数组** 同上结构。
    - **文件**：`multipart/form-data`，`file` 为 `.csv` 或 `.json`。
    """
    rows: list[tuple[str, str, str | None]] = []
    ct = (request.headers.get("content-type") or "").lower()
    if "application/json" in ct:
        try:
            raw = await request.json()
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="invalid json",
            ) from e
        rows = _rows_from_json_raw(raw)
    elif "multipart/form-data" in ct:
        if not file or not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="file required in multipart",
            )
        raw = await file.read()
        rows = _load_rows_from_file(raw, file.filename)
    else:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="use application/json or multipart/form-data with file",
        )
    n = 0
    for sn, fn, class_label in rows:
        if not sn or not fn:
            continue
        cid = await _get_or_create_class_id(db, class_label)
        await _upsert_roster_row(db, sn, fn, cid)
        n += 1
    return {"ok": True, "imported": n}


# --- 任务 5：教师凭 **再验管理员密码** 查看学生存储的可逆口令；写 `admin_audit`（设计 §3.4）---


class RevealPasswordBody(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    admin_password: str = Field(
        min_length=1,
        description="当前管理员登录密码；用于防误读/越权展示。",
        alias="adminPassword",
    )


@router.post(
    "/students/{student_id}/reveal-password",
    summary="解密并返回学生登录密码（课堂排障，高敏感）",
)
async def reveal_student_password(
    _t: CurrentTeacher,
    student_id: uuid.UUID,
    body: RevealPasswordBody,
    db: DBSession,
    request: Request,
) -> dict:
    """
    **高敏感**：解密 `students.password_ciphertext` 并返回明文，**仅**用于课堂内协助学生排障。

    须带 `teacher_session`，且 body 中 **`adminPassword`** 与当前管理员口令一致；成功时写入审计表 `admin_audit`（`view_student_password`）。

    设计 §3.4 要求再验管理员密码；本实现为 **POST** 与 OpenAPI 描述符一致（避免 GET 带 body 的互操作问题）。
    """
    r = await db.execute(select(AdminConfig).where(AdminConfig.id == 1))
    admin = r.scalar_one_or_none()
    if admin is None or _pwd is None or not _pwd.verify(body.admin_password, admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="invalid admin password",
        )
    sr = await db.execute(select(Student).where(Student.id == student_id))
    st = sr.scalar_one_or_none()
    if st is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="student not found")
    try:
        plain = decrypt_password(st.password_ciphertext)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="decrypt failed",
        ) from e
    client = request.client
    ip = client.host if client else None
    ua = request.headers.get("user-agent")
    db.add(
        AdminAudit(
            action="view_student_password",
            target_student_id=st.id,
            ip=ip[:64] if ip else None,
            user_agent=ua[:512] if ua else None,
        )
    )
    return {
        "ok": True,
        "studentId": str(st.id),
        "studentNo": st.student_no,
        "password": plain,
        "auditedAt": datetime.now(timezone.utc).isoformat(),
    }
